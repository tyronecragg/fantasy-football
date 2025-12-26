import re
import csv
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def convert_special_characters(text):
    """
    Convert special characters to their English alphabet equivalents
    """
    char_map = {
        'Đ': 'Dj', 'đ': 'dj',
        'á': 'a', 'à': 'a', 'ä': 'a', 'â': 'a', 'ā': 'a', 'ã': 'a', 'å': 'a',
        'Á': 'A', 'À': 'A', 'Ä': 'A', 'Â': 'A', 'Ā': 'A', 'Ã': 'A', 'Å': 'A',
        'é': 'e', 'è': 'e', 'ë': 'e', 'ê': 'e', 'ē': 'e',
        'É': 'E', 'È': 'E', 'Ë': 'E', 'Ê': 'E', 'Ē': 'E',
        'í': 'i', 'ì': 'i', 'ï': 'i', 'î': 'i', 'ī': 'i',
        'Í': 'I', 'Ì': 'I', 'Ï': 'I', 'Î': 'I', 'Ī': 'I',
        'ó': 'o', 'ò': 'o', 'ö': 'o', 'ô': 'o', 'ō': 'o', 'õ': 'o', 'ø': 'o',
        'Ó': 'O', 'Ò': 'O', 'Ö': 'O', 'Ô': 'O', 'Ō': 'O', 'Õ': 'O', 'Ø': 'O',
        'ú': 'u', 'ù': 'u', 'ü': 'u', 'û': 'u', 'ū': 'u',
        'Ú': 'U', 'Ù': 'U', 'Ü': 'U', 'Û': 'U', 'Ū': 'U',
        'ñ': 'n', 'Ñ': 'N',
        'ç': 'c', 'Ç': 'C',
        'ý': 'y', 'ÿ': 'y', 'Ý': 'Y', 'Ÿ': 'Y',
        'ž': 'z', 'Ž': 'Z',
        'š': 's', 'Š': 'S',
        'č': 'c', 'Č': 'C',
        'ř': 'r', 'Ř': 'R',
        'ď': 'd', 'Ď': 'D',
        'ť': 't', 'Ť': 'T',
        'ň': 'n', 'Ň': 'N',
        'ľ': 'l', 'Ľ': 'L',
        'ĺ': 'l', 'Ĺ': 'L',
        'ŕ': 'r', 'Ŕ': 'R',
        # 'đ': 'd', 'Đ': 'D',
        'ș': 's', 'Ș': 'S',
        'ț': 't', 'Ț': 'T',
        'ă': 'a', 'Ă': 'A',
        'ğ': 'g', 'Ğ': 'G',
        'ı': 'i', 'İ': 'I',
        'ş': 's', 'Ş': 'S',
        'ć': 'c', 'Ć': 'C',
        'ł': 'l', 'Ł': 'L',
        'ń': 'n', 'Ń': 'N',
        'ś': 's', 'Ś': 'S',
        'ź': 'z', 'Ź': 'Z',
        'ż': 'z', 'Ż': 'Z',
        'ą': 'a', 'Ą': 'A',
        'ę': 'e', 'Ę': 'E',
        'ő': 'o', 'Ő': 'O',
        'ű': 'u', 'Ű': 'U',
        'æ': 'ae', 'Æ': 'AE',
        'œ': 'oe', 'Œ': 'OE',
        'ß': 'ss',
        'Ã¡': 'a', 'Ã©': 'e', 'Ã­': 'i', 'Ã³': 'o', 'Ãº': 'u',
        'Ã±': 'n', 'Ã«': 'e', 'Ã¶': 'o', 'Ã¼': 'u',
        "’": "'",
    }

    # Replace special characters
    result = text
    for special_char, replacement in char_map.items():
        result = result.replace(special_char, replacement)

    return result


def normalise_player_name(name):

    name = convert_special_characters(name)

    name_mappings = {
        "David Raya Martin": "David Raya",
        "Matthew Cash": "Matty Cash",
        "Dorde Petrovic": "Djordje Petrovic",
        "Ezri Konsa Ngoyo": "Ezri Konsa",
        "Tosin": "Tosin Adarabioyo",
        "Alisson": "Alisson Becker",
        "Toti": "Toti Gomes",
        "Jan-Paul van Hecke": "Jan Paul van Hecke",
        "Ehor Yarmolyuk": "Yehor Yarmoliuk",
        "Nico Oâ€™Reilly": "Nico O'Reilly",
        "Morato": "Felipe Rodrigues da Silva",
        "Danny Ballard": "Dan Ballard",
        "Ladislav Krejci II": "Ladislav Krejci",
        "Thiago": "Igor Thiago",
        "Dalot": "Diogo Dalot",
        "Hugo Bueno": "Santiago Bueno",
        "Yeremi Pino": "Yeremy Pino",
        "Vitaliy Mykolenko": "Vitalii Mykolenko",
        "Joshua King": "Josh King",
        "Valentino Livramento": "Tino Livramento",
        "Max Kilman": "Maximilian Kilman",
        "El Hadji Diouf": "El Hadji Malick Diouf",
        "Lamar Bogarde": "Lamare Bogarde",
        "Garnacho": "Alejandro Garnacho",
        "Murillo": "Murillo dos Santos",
        "Gabriel": "Gabriel Magalhaes",
        "Amadou Mvom Onana": "Amadou Onana",
        "Murillo": "Murillo dos Santos",
        "Reinildo": "Reinildo Mandava",
        "Pape Sarr": "Pape Matar Sarr",
        "Emile Smith-Rowe": "Emile Smith Rowe",
        "Joseph Gomez": "Joe Gomez",
        "Joseph Willock": "Joe Willock",
        "Lindelof": "Victor Lindelof",
        "Murillo dos Santos": "Murillo",
    }

    # Return mapped name if exists, otherwise return original
    return name_mappings.get(name, name)


def extract_full_name(title_text):
    if not title_text:
        return ""

    # Check if there's a pattern like "Last Name (First Name)"
    bracket_match = re.search(r'^(.+?)\s*\((.+?)\)$', title_text)
    if bracket_match:
        last_part = bracket_match.group(1).strip()
        first_name = bracket_match.group(2).strip()
        name = f"{first_name} {last_part}"
        name_normalised = normalise_player_name(name)
    else:
        name = title_text.strip()
        name_normalised = normalise_player_name(name)
    return name_normalised


def get_team_lineups():
    """
    Scrape FFS team news page for predicted lineups
    """
    url = "https://www.fantasyfootballscout.co.uk/team-news"

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        teams_data = []

        # Find all team news items
        team_items = soup.find_all('li', class_='team-news-item')

        for team_item in team_items:
            # Get team name
            team_header = team_item.find('header')
            if not team_header:
                continue

            team_name_elem = team_header.find('h2')
            if not team_name_elem:
                continue

            team_name = team_name_elem.get_text(strip=True)

            # Find the formation/lineup section
            formation_div = team_item.find('div', class_='scout-picks-pitch')
            if not formation_div:
                continue

            # Get all player elements from all rows (excluding bench)
            player_elements = []

            # Look for rows 1-5 (starting XI, excluding bench row-5 in some formations)
            for row_num in range(1, 6):
                row = formation_div.find('ul', class_=f'row-{row_num}')
                if row:
                    players_in_row = row.find_all('li')
                    for player in players_in_row:
                        # Skip reserve players (bench players)
                        if 'reserve' not in player.get('class', []):
                            player_elements.append(player)

            # Extract player names from starting XI
            for player_elem in player_elements:
                title = player_elem.get('title', '')
                if title:
                    full_name = extract_full_name(title)
                    if full_name:
                        teams_data.append({
                            'Player': full_name,
                            'Team': team_name,
                        })

        return teams_data

    except requests.RequestException as e:
        print(f"Error fetching the webpage: {e}")
        return []
    except Exception as e:
        print(f"Error parsing the webpage: {e}")
        return []


def populate_excel_columns(excel_file, sheet_name, dataframe, start_row=2, include_headers=True):
    """
    Populate Excel columns starting from A with dataframe data and optionally headers.

    Parameters:
    - excel_file: path to Excel file
    - sheet_name: name of the worksheet
    - dataframe: pandas DataFrame with data to insert
    - start_row: row number to start inserting data (default: 2)
    - include_headers: whether to write column names to row 1 (default: True)
    """

    workbook = load_workbook(excel_file)
    worksheet = workbook[sheet_name]

    num_cols = len(dataframe.columns)

    # Add headers if requested
    if include_headers:
        for col_idx, col_name in enumerate(dataframe.columns):
            excel_col = get_column_letter(col_idx + 1)
            worksheet[f'{excel_col}1'].value = col_name

    # Clear existing data in the data range
    for col_idx in range(num_cols):
        excel_col = get_column_letter(col_idx + 1)  # A=1, B=2, etc.
        # Clear the entire column by deleting all values from start_row to max_row
        for row in range(start_row, worksheet.max_row + 1):
            worksheet[f'{excel_col}{row}'].value = None

    # Insert new data
    for row_idx, (_, row_data) in enumerate(dataframe.iterrows(), start=start_row):
        for col_idx, col_name in enumerate(dataframe.columns):
            excel_col = get_column_letter(col_idx + 1)
            value = row_data[col_name]

            # Handle NaN values
            if pd.isna(value):
                worksheet[f'{excel_col}{row_idx}'].value = None
            else:
                worksheet[f'{excel_col}{row_idx}'].value = value

    workbook.save(excel_file)

    print(f"Successfully updated columns A-{get_column_letter(num_cols)} ({num_cols} columns, {len(dataframe)} rows) in {sheet_name}")


def save(teams_data, filename="starting_lineups/data.csv"):
    """
    Save the teams data to CSV file
    """
    if not teams_data:
        print("No data to save")
        return

    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        column_names = ['Player', 'Team']
        writer = csv.DictWriter(csvfile, fieldnames=column_names)

        writer.writeheader()
        for row in teams_data:
            writer.writerow(row)

    df = pd.DataFrame(teams_data, columns=column_names)
    populate_excel_columns('Fantasy Premier League.xlsx', 'Starting Lineups', df)

    print(f"Data saved to {filename}")
    print(f"Total players extracted: {len(teams_data)}")


def main():
    """
    Main function to run the scraper
    """
    print("Scraping FFS Team News for predicted lineups...")

    teams_data = get_team_lineups()

    if teams_data:
        # Display sample data
        print("\nSample of extracted data:")
        for i, player in enumerate(teams_data[:10]):  # Show first 10 players
            print(f"{player['Player']}, {player['Team']}")

        if len(teams_data) > 10:
            print("...")

        # Save to CSV
        save(teams_data)

        # Summary by team
        teams_count = {}
        for player in teams_data:
            team = player['Team']
            teams_count[team] = teams_count.get(team, 0) + 1

        print(f"\nPlayers per team:")
        for team, count in sorted(teams_count.items()):
            print(f"{team}: {count} players")

    else:
        print("No data was extracted.")


if __name__ == "__main__":
    main()
