"""One-off migration: export every manually-maintained sheet from the workbook into
editable CSVs under inputs/. After this, the pipeline never needs the workbook except
for parity validation.

Only raw (hand-entered or orphaned-source) columns are exported — anything the workbook
derives with formulas is recomputed by the pipeline instead.
"""
import os

import openpyxl
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "Fantasy Premier League.xlsx")
INPUTS = os.path.join(ROOT, "inputs")

# sheet -> (output csv, column letters to keep, n header rows assumed 1)
EXPORTS = {
    "Title Odds": ("title_odds.csv", None),          # A + all bookmaker cols B:Y (drop formula col AD)
    "Relegation Odds": ("relegation_odds.csv", None),
    "Top 6 Odds": ("top6_odds.csv", None),
    "Fixtures": ("fixtures.csv", "all"),
    "Starting Lineups": ("starting_lineups.csv", "A:H"),
    "GW Teams": ("gw_teams.csv", "all"),
    "Fallback Factors": ("fallback_factors.csv", "all"),
    "F2 Yellow Card": ("f2_yellow_card.csv", "A:D"),  # source pipeline deleted; keep as editable input
    "Historical Fixture Odds": ("historical_fixture_odds.csv", "A:J"),  # raw odds; percs recomputed
    "Historical Player Data": ("historical_player_data.csv", "all"),
    "Historical Expected Points": ("historical_expected_points.csv", "all"),
}


def sheet_to_df(ws, col_range):
    data = [[c.value for c in row] for row in ws.iter_rows()]
    df = pd.DataFrame(data[1:], columns=data[0])
    if col_range == "all" or col_range is None:
        pass
    else:
        start, end = col_range.split(":")
        i0 = openpyxl.utils.column_index_from_string(start) - 1
        i1 = openpyxl.utils.column_index_from_string(end)
        df = df.iloc[:, i0:i1]
    return df.dropna(how="all")


def main():
    os.makedirs(INPUTS, exist_ok=True)
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    for sheet, (csv_name, col_range) in EXPORTS.items():
        ws = wb[sheet]
        df = sheet_to_df(ws, col_range)
        if sheet in ("Title Odds", "Relegation Odds", "Top 6 Odds"):
            # A = team, B:Y = individual bookmaker odds; AD is a formula (recomputed)
            df = df.iloc[:, 0:25]
            df.columns = ["Team"] + [f"book_{i}" for i in range(1, len(df.columns))]
        out = os.path.join(INPUTS, csv_name)
        df.to_csv(out, index=False)
        print(f"{sheet:<28} -> inputs/{csv_name:<34} {df.shape[0]} rows x {df.shape[1]} cols")

    # Defensive-contribution distribution parameters (manual cells G2/H2 on each sheet)
    rows = []
    for pos, sheet in [("DEF", "Defensive Contribution DEF"), ("MID", "Defensive Contribution MID")]:
        ws = wb[sheet]
        rows.append({"position": pos, "threshold": 10 if pos == "DEF" else 12,
                     "sd": ws["G2"].value, "average_dc90": ws["H2"].value})
    pd.DataFrame(rows).to_csv(os.path.join(INPUTS, "dc_params.csv"), index=False)
    print(f"{'DC G2/H2 cells':<28} -> inputs/dc_params.csv")


if __name__ == "__main__":
    main()
