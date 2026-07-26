"""One-off: extract every hardcoded regression coefficient from the workbook formulas
into fpl_pipeline/data/coefficients.json.

Parsing is intentionally strict: any term that doesn't normalise to a known feature
raises, so a formula change in the workbook can't silently produce wrong coefficients.
"""
import json
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "Fantasy Premier League.xlsx")
OUT = os.path.join(ROOT, "fpl_pipeline", "data", "coefficients.json")


def matching_paren_span(s, open_idx):
    """Index just past the ')' matching the '(' at open_idx."""
    depth = 0
    for i in range(open_idx, len(s)):
        if s[i] == "(":
            depth += 1
        elif s[i] == ")":
            depth -= 1
            if depth == 0:
                return i
    raise ValueError("unbalanced parens")


def split_top_level_terms(expr):
    """Split 'a+b-c' into [('+', 'a'), ('+', 'b'), ('-', 'c')] at paren depth 0."""
    terms, depth, start, sign = [], 0, 0, "+"
    for i, ch in enumerate(expr):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch in "+-" and depth == 0 and i > start:
            terms.append((sign, expr[start:i]))
            sign, start = ch, i + 1
        elif ch in "+-" and depth == 0 and i == start:
            sign = ch  # leading sign
            start = i + 1
    terms.append((sign, expr[start:]))
    return [(s, t) for s, t in terms if t]


# --- player-stat baselines: features over win prob / opponent win prob / position / venue ---

def normalise_baseline_term(term, win, opp, venue):
    """Map one formula term to (feature, coefficient)."""
    t = term.replace(" ", "").replace("$", "")
    t = t.replace(f'IF(B2="DEF",1,0)', "DEF").replace(f'IF(B2="MID",1,0)', "MID")
    t = t.replace(f'IF(B2="FWD",1,0)', "FWD").replace(f'IF({venue}="H",1,0)', "HOME")
    t = t.replace(f"ABS({win}-{opp})", "ABSDIFF")
    t = t.replace(f"({win}-{opp})^2", "DIFF2")
    t = t.replace(f"({win}-{opp})", "DIFFP")  # parenthesised diff
    t = t.replace(win, "WIN").replace(opp, "OPP")

    m = re.match(r"^(\d+\.?\d*)(?:\*(.+))?$", t)
    if not m:
        raise ValueError(f"unparseable term: {term!r} -> {t!r}")
    coef, expr = float(m.group(1)), (m.group(2) or "const")

    feature_map = {
        "const": "const",
        "WIN": "win", "OPP": "opp",
        "DEF": "def", "MID": "mid", "FWD": "fwd",
        "HOME": "home",
        "DIFFP": "diff", "ABSDIFF": "absdiff", "DIFF2": "diff2",
        "(WIN*HOME)": "win_home", "(OPP*HOME)": "opp_home",
        "(DIFFP*HOME)": "diff_home",
        "(DEF*DIFFP)": "def_diff", "(MID*DIFFP)": "mid_diff", "(FWD*DIFFP)": "fwd_diff",
        "(DEF*HOME)": "def_home", "(MID*HOME)": "mid_home", "(FWD*HOME)": "fwd_home",
        "(WIN*OPP)": "win_opp",
    }
    if expr not in feature_map:
        raise ValueError(f"unknown feature {expr!r} from term {term!r}")
    return feature_map[expr], coef


def parse_baseline(expr, win="I2", opp="J2", venue="M2"):
    out = {}
    for sign, term in split_top_level_terms(expr):
        feat, coef = normalise_baseline_term(term, win, opp, venue)
        assert feat not in out, f"duplicate feature {feat}"
        out[feat] = coef if sign == "+" else -coef
    return out


def factor_denominator(formula, numerator):
    """From =IF(NOT(ISNA(I2)),NUM/(EXPR),VLOOKUP(...)) pull EXPR."""
    key = f"{numerator}/("
    i = formula.index(key) + len(key) - 1
    j = matching_paren_span(formula, i)
    return formula[i + 1:j]


# --- F3-F6 win prediction: features over season-odds probabilities ---

def parse_win_pred(formula):
    expr = formula.lstrip("=").replace(" ", "").replace("$", "")
    # canonical replacements, longest first (E2=title, F2=releg, G2=top6; BP/BQ/BR opponent)
    strdiff = "((E2+G2-F2)-(BP2+BR2-BQ2))"
    repl = [
        (f"ABS({strdiff})", "ABSSD"),
        (f"({strdiff}^2)", "SD2"),
        (f"(IF(BO2=\"H\",1,0)*{strdiff})", "HOMESD"),
        (strdiff, "SD"),
        ('IF(BO2="H",1,0)', "HOME"),
        ("(G2/(G2+BR2+0.01))", "TOP6SHARE"),
        ("(E2-BP2)", "TITLEDIFF"),
        ("BR2", "OPPTOP6"),
        ("E2", "TITLE"),
    ]
    for a, b in repl:
        expr = expr.replace(a, b)
    feature_map = {"const": "const", "SD": "strength_diff", "HOME": "home",
                   "TOP6SHARE": "top6_share", "TITLEDIFF": "title_diff",
                   "HOMESD": "home_x_strength_diff", "OPPTOP6": "opp_top6",
                   "TITLE": "title", "SD2": "strength_diff_sq", "ABSSD": "abs_strength_diff"}
    out = {}
    for sign, term in split_top_level_terms(expr):
        m = re.match(r"^(\d+\.?\d*)(?:\*(.+))?$", term)
        if not m:
            raise ValueError(f"unparseable win-pred term {term!r}")
        coef, e = float(m.group(1)), (m.group(2) or "const")
        e = e.strip("()") if e.startswith("(") and e.endswith(")") and "*" not in e else e
        if e not in feature_map:
            raise ValueError(f"unknown win-pred feature {e!r}")
        out[feature_map[e]] = coef if sign == "+" else -coef
    return out


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    ws = wb["Players"]

    baselines = {}
    for stat, cell, num in [
        ("score1", "AC2", "Q2"), ("assist", "AD2", "T2"), ("yellow", "AE2", "U2"),
        ("concede2", "AF2", "W2"), ("concede4", "AG2", "X2"),
        ("saves3", "AH2", "Y2"), ("saves6", "AI2", "Z2"),
    ]:
        baselines[stat] = parse_baseline(factor_denominator(ws[cell].value, num))

    # Clean sheet baseline lives in BI2 as =$V2/(EXPR)
    bi = ws["BI2"].value.replace("=$V2/(", "", 1)
    baselines["clean_sheet"] = parse_baseline(bi[: bi.rfind(")")])

    # F1 Pred XP (col AM): whole-expression regression, same feature space
    baselines["pred_xp"] = parse_baseline(ws["AM2"].value.lstrip("="))

    win_pred = parse_win_pred(ws["BS2"].value)

    # Bonus model: =MAX(0, MIN(1, a + b * AJ2))
    m = re.search(r"MIN\(1,\s*(-?\d+\.?\d*)\s*\+\s*(-?\d+\.?\d*)\s*\*", ws["AK2"].value)
    bonus = {"intercept": float(m.group(1)), "slope": float(m.group(2))}

    cs = wb["Coefficients"]
    sheet_coefs = {cs.cell(row=r, column=1).value: cs.cell(row=r, column=2).value
                   for r in range(1, 47)}

    out = {
        "_source": "Extracted from 'Fantasy Premier League.xlsx' formulas by tools/extract_coefficients.py",
        "baselines": baselines,
        "win_pred_f3plus": win_pred,
        "bonus": bonus,
        "coefficients_sheet": sheet_coefs,
        "total_xp_weights": [1.0, 0.85, 0.7, 0.7, 0.7, 0.7],
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(out, fh, indent=1)
    print(f"Wrote {OUT}")
    for stat, feats in baselines.items():
        print(f"  {stat:<12} {len(feats)} terms")
    print(f"  win_pred     {len(win_pred)} terms")


if __name__ == "__main__":
    main()
