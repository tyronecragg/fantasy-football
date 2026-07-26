"""Parity harness: compare the pipeline's master DataFrame against the workbook's
cached Players values (what Excel last calculated).

Excel error strings (#N/A etc.) are treated as NaN — the pipeline models them as NaN.
"""
import numpy as np
import openpyxl
import pandas as pd

from . import config
from .io_utils import snapshot

TOL = 1e-8


def load_workbook_players():
    wb = openpyxl.load_workbook(config.WORKBOOK, data_only=True, read_only=True)
    ws = wb["Players"]
    rows = list(ws.values)
    wb.close()
    df = pd.DataFrame(rows[1:], columns=rows[0])
    return df.replace(r"^#.*", np.nan, regex=True)


def run(master):
    excel = load_workbook_players()

    # Row alignment sanity check: both should derive from the same roster in the same order
    n = min(len(master), len(excel))
    name_mismatch = (master["Player Name"].iloc[:n].values != excel["Player Name"].iloc[:n].values).sum()
    print(f"\nParity: {len(master)} pipeline rows vs {len(excel)} sheet rows; "
          f"{name_mismatch} name mismatches in aligned rows")

    records = []
    for col in master.columns:
        if col not in excel.columns:
            records.append({"column": col, "status": "missing in sheet"})
            continue
        mine = master[col].iloc[:n].reset_index(drop=True)
        theirs = excel[col].iloc[:n].reset_index(drop=True)

        # Excel VLOOKUP renders a found-but-blank cell as 0 in passthrough lookup columns
        # (Opponent / Venue); the pipeline keeps NaN there. Treat that pattern as equal.
        lookup_col = col.endswith(" Opponent") or col.endswith(" Venue")

        if pd.api.types.is_numeric_dtype(mine):
            theirs_num = pd.to_numeric(theirs, errors="coerce")
            both = mine.notna() & theirs_num.notna()
            close = np.isclose(mine.fillna(0), theirs_num.fillna(0), atol=TOL, rtol=1e-9)
            mismatch = (~close & both) | (mine.isna() != theirs_num.isna())
            if lookup_col:
                mismatch &= ~(mine.isna() & (theirs_num == 0))
            diffs = (mine[both] - theirs_num[both]).abs()
            max_diff = float(diffs.max()) if len(diffs) else 0.0
        else:
            mine_s = mine.astype(object).where(mine.notna())
            theirs_s = theirs.astype(object).where(theirs.notna())
            blank_as_zero = mine_s.isna() & (theirs_s == 0) if lookup_col else False
            mismatch = ~((mine_s == theirs_s) | (mine_s.isna() & theirs_s.isna()) | blank_as_zero)
            max_diff = np.nan

        n_bad = int(mismatch.sum())
        rec = {"column": col, "status": "ok" if n_bad == 0 else "MISMATCH",
               "n_mismatch": n_bad, "max_abs_diff": max_diff}
        if n_bad:
            i = mismatch[mismatch].index[0]
            rec["first_row"] = i + 2  # sheet row number
            rec["first_player"] = master["Player Name"].iloc[i]
            rec["mine"] = master[col].iloc[i]
            rec["excel"] = excel[col].iloc[i]
        records.append(rec)

    report = pd.DataFrame(records)
    snapshot(report, "parity_report")

    bad = report[report["status"] == "MISMATCH"]
    print(f"Columns compared: {len(report)}   clean: {(report['status'] == 'ok').sum()}   "
          f"mismatching: {len(bad)}")
    if len(bad):
        print("\nWorst columns:")
        cols = ["column", "n_mismatch", "max_abs_diff", "first_row", "first_player", "mine", "excel"]
        print(bad.sort_values("n_mismatch", ascending=False)[cols].head(20).to_string(index=False))
    return report
