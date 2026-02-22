import pandas as pd
import os
from openpyxl import load_workbook

SEASON = "2025-2026"
REPO_PATH = os.path.join("fpl_data", "FPL-Core-Insights")
NAME_CHANGES_PATH = "fpl_data"
DATA_PATH = os.path.join(REPO_PATH, "data", SEASON)
EXCEL_FILE = "Fantasy Premier League.xlsx"
SHEET_NAME = "FPL Players"

POSITION_MAP = {
    "Goalkeeper": "GK",
    "Defender": "DEF",
    "Midfielder": "MID",
    "Forward": "FWD",
}


def main():
    players = pd.read_csv(os.path.join(DATA_PATH, "players.csv"))
    player_stats = pd.read_csv(os.path.join(DATA_PATH, "playerstats.csv"))
    teams = pd.read_csv(os.path.join(DATA_PATH, "teams.csv"))

    team_map = teams.set_index("code")["name"].to_dict()
    players["team"] = players["team_code"].map(team_map)
    players["position"] = players["position"].map(POSITION_MAP)

    player_stats = player_stats.sort_values("gw", ascending=True).drop_duplicates(subset="id", keep="last")
    player_stats = player_stats[["id", "now_cost"]]

    df = players.merge(player_stats, left_on="player_id", right_on="id", how="left")

    df["name"] = df["first_name"] + " " + df["second_name"]

    name_changes_path = os.path.join(NAME_CHANGES_PATH, "player_name_changes.csv")
    name_changes = pd.read_csv(name_changes_path)
    name_map = dict(zip(name_changes["name"], name_changes["name_cleaned"]))
    df["name"] = df["name"].map(name_map).fillna(df["name"])

    df = df.rename(columns={"now_cost": "cost"})

    pos_order = pd.CategoricalDtype(categories=["GK", "DEF", "MID", "FWD"], ordered=True)
    df["position"] = df["position"].astype(pos_order)

    output = (
        df[["name", "position", "team", "cost"]]
        .sort_values(["team", "position", "name", "cost"], ascending=[True, True, True, False])
        .reset_index(drop=True)
    )

    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_FILE)
    wb = load_workbook(excel_path)

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_NAME)

    # Header
    headers = ["name", "position", "team", "cost"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Data
    for row_idx, row in enumerate(output.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(excel_path)
    print(f"Wrote {len(output)} players to '{SHEET_NAME}' in {EXCEL_FILE}")


if __name__ == "__main__":
    main()
