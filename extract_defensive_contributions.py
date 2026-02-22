import pandas as pd
import os
from openpyxl import load_workbook

SEASON = "2025-2026"
REPO_PATH = os.path.join("fpl_data", "FPL-Core-Insights")
NAME_CHANGES_PATH = "fpl_data"
DATA_PATH = os.path.join(REPO_PATH, "data", SEASON)
EXCEL_FILE = "Fantasy Premier League.xlsx"

POSITION_MAP = {
    "Goalkeeper": "GK",
    "Defender": "DEF",
    "Midfielder": "MID",
    "Forward": "FWD",
}


def write_sheet(wb, sheet_name, df):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_col=1, max_col=3, min_row=2, max_row=ws.max_row+200):
            for cell in row:
                cell.value = None

    # headers = list(df.columns)
    # for col, header in enumerate(headers, 1):
    #     ws.cell(row=1, column=col, value=header)

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def main():
    players = pd.read_csv(os.path.join(DATA_PATH, "players.csv"))
    player_stats = pd.read_csv(os.path.join(DATA_PATH, "playerstats.csv"))

    players["position"] = players["position"].map(POSITION_MAP)

    player_stats = player_stats.sort_values("gw", ascending=True).drop_duplicates(subset="id", keep="last")
    player_stats = player_stats[["id", "minutes", "defensive_contribution_per_90"]]

    df = players.merge(player_stats, left_on="player_id", right_on="id", how="left")

    df["name"] = df["first_name"] + " " + df["second_name"]

    name_changes_path = os.path.join(NAME_CHANGES_PATH, "player_name_changes.csv")
    name_changes = pd.read_csv(name_changes_path)
    name_map = dict(zip(name_changes["name"], name_changes["name_cleaned"]))
    df["name"] = df["name"].map(name_map).fillna(df["name"])

    df["90s"] = (df["minutes"] / 90).round(2)
    df = df.rename(columns={"defensive_contribution_per_90": "dc_per_90"})

    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_FILE)
    wb = load_workbook(excel_path)

    for position, sheet_name in [("DEF", "Defensive Contribution DEF"), ("MID", "Defensive Contribution MID")]:
        output = (
            df[df["position"] == position][["name", "90s", "dc_per_90"]]
            .sort_values("name")
            .reset_index(drop=True)
        )
        write_sheet(wb, sheet_name, output)
        print(f"Wrote {len(output)} {position} players to '{sheet_name}'")

    wb.save(excel_path)
    print(f"Saved {EXCEL_FILE}")


if __name__ == "__main__":
    main()
