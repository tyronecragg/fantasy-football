import json
import time
import requests
import pandas as pd
from typing import List, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ComprehensiveSportsbetScraper:
    def __init__(self):
        self.base_url = "https://www.sportsbet.com.au/apigw/sportsbook-sports"
        self.session = requests.Session()
        # Add headers to mimic a browser request
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:144.0) Gecko/20100101 Firefox/144.0',
            'Accept': 'application/json',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })

    def format_datetime(self, datetime_str: str) -> str:
        try:
            if not datetime_str:
                return ""
            # Parse the datetime (adjust format based on actual API response)
            dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
            return dt.strftime('%a %d %b %H:%M')
        except:
            return datetime_str

    def get_match_ids(self, competition_id: int = 718) -> List[Dict]:
        """
        Get all match IDs for the Premier League (competition ID 718)
        """
        url = f"{self.base_url}/Sportsbook/Sports/Competitions/{competition_id}"
        params = {
            'displayType': 'default',
            'includeTopMarkets': 'true',
            'eventFilter': 'matches'
        }

        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            data = response.json()

            matches = []
            if 'events' in data:
                for event in data['events']:
                    participant1 = event.get('participant1')
                    participant2 = event.get('participant2')

                    if participant1 == 'Tottenham':
                        participant1 = 'Spurs'
                    if participant2 == 'Tottenham':
                        participant2 = 'Spurs'

                    if participant1 == 'Nottm Forest':
                        participant1 = "Nott'm Forest"
                    if participant2 == 'Nottm Forest':
                        participant2 = "Nott'm Forest"

                    matches.append({
                        'match_id': event['id'],
                        'match_name': event['name'],
                        'start_time': event.get('startTime'),
                        'formatted_date': self.format_datetime(event.get('startTime')),
                        'participant1': participant1,
                        'participant2': participant2,
                    })

            print(f"Found {len(matches)} matches")
            return matches

        except requests.exceptions.RequestException as e:
            print(f"Error fetching match IDs: {e}")
            return []
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response: {e}")
            return []

    def discover_markets(self, match_id: int) -> Dict:
        """
        Get all markets for a match to discover market IDs
        """
        url = f"{self.base_url}/Sportsbook/Sports/Events/{match_id}/Markets"

        try:
            response = self.session.get(url)
            response.raise_for_status()
            data = response.json()

            print(f"\nAvailable markets for match {match_id}:")
            if isinstance(data, list):
                for market in data:
                    market_id = market.get('id')
                    market_name = market.get('name')
                    print(f"  ID: {market_id} - Name: {market_name}")

            return data

        except Exception as e:
            print(f"Error discovering markets: {e}")
            return {}

    def get_match_markets(self, match_id: int, market_ids: str = None) -> Dict:
        """
        Get specific markets for a match
        """
        url = f"{self.base_url}/Sportsbook/Sports/Events/{match_id}/Markets"
        params = {}

        if market_ids:
            params['marketIds'] = market_ids

        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            response_data = response.json()
            return response_data

        except requests.exceptions.RequestException as e:
            print(f"Error fetching markets for match {match_id}: {e}")
            return {}
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response for match {match_id}: {e}")
            return {}

    def extract_win_draw_win_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract Win-Draw-Win (1X2) match result odds
        """
        win_draw_win_odds = []

        if not isinstance(markets_data, list):
            return win_draw_win_odds

        for market in markets_data:
            market_name = market.get('name', '')
            if any(keyword in market_name for keyword in ['Win-Draw-Win']):

                home_win = away_win = draw = None

                for selection in market.get('selections', []):
                    selection_name = selection.get('name', '')
                    odds = selection.get('price', {}).get('winPrice')

                    # Normalize selection name for comparison
                    selection_name_normalized = selection_name
                    if selection_name == "Tottenham":
                        selection_name_normalized = "Spurs"
                    if selection_name == "Nottm Forest":
                        selection_name_normalized = "Nott'm Forest"

                    # Now compare using normalized names
                    if match_info.get('participant1', '').lower() in selection_name_normalized.lower():
                        home_win = odds
                    elif match_info.get('participant2', '').lower() in selection_name_normalized.lower():
                        away_win = odds
                    elif 'draw' in selection_name.lower():
                        draw = odds

                if home_win or away_win or draw:
                    win_draw_win_odds.append({
                        'home_team': match_info.get('participant1'),
                        'away_team': match_info.get('participant2'),
                        'home_win_odds': home_win,
                        'away_win_odds': away_win,
                    })
                break

        return win_draw_win_odds

    def extract_goalscorer_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract anytime goalscorer odds from markets data
        """
        goalscorer_odds = []

        if not isinstance(markets_data, list):
            return goalscorer_odds

        for market in markets_data:
            market_name = market.get('name', '')
            # Look for anytime goalscorer markets
            if 'Anytime Goalscorer' in market_name or 'Goal Scorer' in market_name:
                for selection in market.get('selections', []):
                    player_odds = {
                        'player_name': selection.get('name'),
                        'match_id': match_info['match_id'],
                        'odds_decimal': selection.get('price', {}).get('winPrice'),
                        # 'match_name': match_info['match_name'],
                        # 'date': match_info.get('formatted_date', ''),
                        # 'home_team': match_info.get('participant1'),
                        # 'away_team': match_info.get('participant2'),
                        # 'market_type': 'Anytime Goalscorer',
                        # 'market_id': market.get('id'),
                        # 'selection_id': selection.get('id'),
                    }
                    goalscorer_odds.append(player_odds)
                break

        return goalscorer_odds

    def extract_assist_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract anytime assist odds from markets data
        """
        assist_odds = []

        if not isinstance(markets_data, list):
            return assist_odds

        for market in markets_data:
            market_name = market.get('name', '')
            # Look for anytime assist markets
            if 'Anytime Assist' in market_name or 'To Assist' in market_name:
                for selection in market.get('selections', []):
                    player_odds = {
                        'player_name': selection.get('name'),
                        'match_id': match_info['match_id'],
                        'odds_decimal': selection.get('price', {}).get('winPrice'),
                        # 'match_name': match_info['match_name'],
                        # 'date': match_info.get('formatted_date', ''),
                        # 'home_team': match_info.get('participant1'),
                        # 'away_team': match_info.get('participant2'),
                        # 'market_type': 'Anytime Assist',
                        # 'market_id': market.get('id'),
                        # 'selection_id': selection.get('id'),
                    }
                    assist_odds.append(player_odds)
                break

        return assist_odds

    def extract_to_be_booked_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract 'To Be Booked' (yellow/red card) odds from markets data
        """
        booking_odds = []

        if not isinstance(markets_data, list):
            return booking_odds

        for market in markets_data:
            market_name = market.get('name', '')
            # Look for booking markets
            if any(keyword in market_name for keyword in
                   ['To Be Booked']):
                for selection in market.get('selections', []):
                    player_odds = {
                        'match_name': match_info['match_name'],
                        'date': match_info.get('formatted_date', ''),
                        'player_name': selection.get('name'),
                        'odds_decimal': selection.get('price', {}).get('winPrice'),
                        # 'match_id': match_info['match_id'],
                        # 'home_team': match_info.get('participant1'),
                        # 'away_team': match_info.get('participant2'),
                        # 'market_type': 'To Be Booked',
                        # 'market_id': market.get('id'),
                        # 'market_name': market_name,
                        # 'selection_id': selection.get('id'),
                    }
                    booking_odds.append(player_odds)

        return booking_odds

    def extract_two_or_more_goals_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract 'To Score 2 or More Goals' odds from markets data
        """
        two_goals_odds = []

        if not isinstance(markets_data, list):
            return two_goals_odds

        for market in markets_data:
            market_name = market.get('name', '')
            # Look for 2+ goals markets
            if any(keyword in market_name for keyword in ['2 or More Goals']):
                for selection in market.get('selections', []):
                    player_odds = {
                        'player_name': selection.get('name'),
                        'match_id': match_info['match_id'],
                        'odds_decimal': selection.get('price', {}).get('winPrice'),
                        # 'match_name': match_info['match_name'],
                        # 'date': match_info.get('formatted_date', ''),
                        # 'home_team': match_info.get('participant1'),
                        # 'away_team': match_info.get('participant2'),
                        # 'market_type': 'To Score 2+ Goals',
                        # 'market_id': market.get('id'),
                        # 'market_name': market_name,
                        # 'selection_id': selection.get('id'),
                    }
                    two_goals_odds.append(player_odds)

        return two_goals_odds

    def extract_two_or_more_assists_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract 'Player To Have 2 Or More Assists' odds from markets data
        """
        two_assists_odds = []

        if not isinstance(markets_data, list):
            return two_assists_odds

        for market in markets_data:
            market_name = market.get('name', '')
            # Look for 2+ assists markets
            if any(keyword in market_name for keyword in
                   ['2 Or More Assists']):
                for selection in market.get('selections', []):
                    player_odds = {
                        'player_name': selection.get('name'),
                        'match_id': match_info['match_id'],
                        'odds_decimal': selection.get('price', {}).get('winPrice'),
                        # 'match_name': match_info['match_name'],
                        # 'date': match_info.get('formatted_date', ''),
                        # 'home_team': match_info.get('participant1'),
                        # 'away_team': match_info.get('participant2'),
                        # 'market_type': 'To Have 2+ Assists',
                        # 'market_id': market.get('id'),
                        # 'market_name': market_name,
                        # 'selection_id': selection.get('id'),
                    }
                    two_assists_odds.append(player_odds)

        return two_assists_odds

    def extract_team_goals_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract team-specific Over/Under 1.5 and 3.5 goals odds from markets data
        Returns data in format: one row per team per match with all their odds
        """
        team_goals_odds = []

        if not isinstance(markets_data, list):
            return team_goals_odds

        participant1 = match_info.get('participant1', '')
        participant2 = match_info.get('participant2', '')

        # Initialize odds storage for both teams
        team1_odds = {
            'over_1_5': None, 'under_1_5': None,
            'over_3_5': None, 'under_3_5': None
        }
        team2_odds = {
            'over_1_5': None, 'under_1_5': None,
            'over_3_5': None, 'under_3_5': None
        }

        # Look through all markets to find team-specific over/under goals
        for market in markets_data:
            market_name = market.get('name', '')

            # Check if this is a team-specific over/under goals market
            is_team1_market = False
            is_team2_market = False
            is_1_5_market = False
            is_3_5_market = False

            # Determine which team and which threshold
            if any(keyword in market_name for keyword in ['Home Team', participant1]):
                is_team1_market = True
            elif any(keyword in market_name for keyword in ['Away Team', participant2]):
                is_team2_market = True

            if '1.5' in market_name and 'Over/Under' in market_name and 'Half' not in market_name:
                is_1_5_market = True
            elif '3.5' in market_name and 'Over/Under' in market_name and 'Half' not in market_name:
                is_3_5_market = True

            # Extract odds if this is a relevant market
            if (is_team1_market or is_team2_market) and (is_1_5_market or is_3_5_market):
                for selection in market.get('selections', []):
                    selection_name = selection.get('name', '').lower()
                    odds = selection.get('price', {}).get('winPrice')

                    if odds is None:
                        continue

                    # Determine if this is over or under
                    is_over = 'over' in selection_name
                    is_under = 'under' in selection_name

                    # Store the odds in the appropriate place
                    if is_team1_market:
                        if is_1_5_market and is_over:
                            team1_odds['over_1_5'] = odds
                        elif is_1_5_market and is_under:
                            team1_odds['under_1_5'] = odds
                        elif is_3_5_market and is_over:
                            team1_odds['over_3_5'] = odds
                        elif is_3_5_market and is_under:
                            team1_odds['under_3_5'] = odds

                    elif is_team2_market:
                        if is_1_5_market and is_over:
                            team2_odds['over_1_5'] = odds
                        elif is_1_5_market and is_under:
                            team2_odds['under_1_5'] = odds
                        elif is_3_5_market and is_over:
                            team2_odds['over_3_5'] = odds
                        elif is_3_5_market and is_under:
                            team2_odds['under_3_5'] = odds

        # Create rows for both teams if we have any odds
        if any(team1_odds.values()) or any(team2_odds.values()):
            # Row for Team 1
            team1_row = {
                'Match': f"{participant1} v {participant2}",
                'Date': match_info.get('formatted_date', ''),
                'Team': participant1,
                'Opponent': participant2,
                'Team_Over_1.5': team1_odds['over_1_5'],
                'Team_Under_1.5': team1_odds['under_1_5'],
                'Team_Over_3.5': team1_odds['over_3_5'],
                'Team_Under_3.5': team1_odds['under_3_5'],
                'Opponent_Concedes_Over_1.5': team1_odds['over_1_5'],
                'Opponent_Concedes_Under_1.5': team1_odds['under_1_5'],
                'Opponent_Concedes_Over_3.5': team1_odds['over_3_5'],
                'Opponent_Concedes_Under_3.5': team1_odds['under_3_5'],
                # 'match_id': match_info['match_id'],
            }
            team_goals_odds.append(team1_row)

            # Row for Team 2
            team2_row = {
                'Match': f"{participant1} v {participant2}",
                'Date': match_info.get('formatted_date', ''),
                'Team': participant2,
                'Opponent': participant1,
                'Team_Over_1.5': team2_odds['over_1_5'],
                'Team_Under_1.5': team2_odds['under_1_5'],
                'Team_Over_3.5': team2_odds['over_3_5'],
                'Team_Under_3.5': team2_odds['under_3_5'],
                'Opponent_Concedes_Over_1.5': team2_odds['over_1_5'],
                'Opponent_Concedes_Under_1.5': team2_odds['under_1_5'],
                'Opponent_Concedes_Over_3.5': team2_odds['over_3_5'],
                'Opponent_Concedes_Under_3.5': team2_odds['under_3_5'],
                # 'match_id': match_info['match_id'],
            }
            team_goals_odds.append(team2_row)

        return team_goals_odds

    def extract_clean_sheet_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract clean sheet odds based on opponent's Over/Under 0.5 Goals markets
        """
        clean_sheet_odds = []

        if not isinstance(markets_data, list):
            return clean_sheet_odds

        participant1 = match_info.get('participant1', '')
        participant2 = match_info.get('participant2', '')

        # Look for team-specific Over/Under 0.5 Goals markets
        team1_over_odds = team1_under_odds = None
        team2_over_odds = team2_under_odds = None

        for market in markets_data:
            market_name = market.get('name', '')

            # Check for team-specific Over/Under 0.5 Goals markets
            if 'Over/Under' in market_name and '0.5' in market_name and 'Half' not in market_name:
                is_team1_market = 'Home Team' in market_name or participant1 in market_name
                is_team2_market = 'Away Team' in market_name or participant2 in market_name

                if is_team1_market or is_team2_market:
                    for selection in market.get('selections', []):
                        selection_name = selection.get('name', '').lower()
                        odds = selection.get('price', {}).get('winPrice')

                        if 'over' in selection_name:
                            if is_team1_market:
                                team1_over_odds = odds
                            else:
                                team2_over_odds = odds
                        elif 'under' in selection_name:
                            if is_team1_market:
                                team1_under_odds = odds
                            else:
                                team2_under_odds = odds

        # Create clean sheet odds for each team
        if team2_under_odds and team2_over_odds:
            clean_sheet_odds.append({
                'match_name': match_info['match_name'],
                'date': match_info.get('formatted_date', ''),
                'team_name': participant1,
                'clean_sheet_yes': team2_under_odds,
                'clean_sheet_no': team2_over_odds,
                # 'match_id': match_info['match_id'],
                # 'opponent': participant2,
            })

        if team1_under_odds and team1_over_odds:
            clean_sheet_odds.append({
                'match_name': match_info['match_name'],
                'date': match_info.get('formatted_date', ''),
                'team_name': participant2,
                'clean_sheet_yes': team1_under_odds,
                'clean_sheet_no': team1_over_odds,
                # 'match_id': match_info['match_id'],
                # 'opponent': participant1,
            })

        return clean_sheet_odds

    def extract_goalkeeper_saves_odds(self, markets_data: Dict, match_info: Dict) -> List[Dict]:
        """
        Extract goalkeeper saves odds (3+ and 6+ saves) and compile into single structure
        Structure: Match, Date, Team, 3+ Saves, 6+ Saves
        """
        goalkeeper_saves_odds = []

        if not isinstance(markets_data, list):
            return goalkeeper_saves_odds

        # Dictionary to store saves odds by goalkeeper
        # Format: {goalkeeper_name: {team: team_name, 3_saves: odds, 6_saves: odds}}
        gk_data = {}

        # First pass: collect all goalkeeper save odds
        for market in markets_data:
            market_name = market.get('name', '')

            # Check for 3+ saves markets
            if any(keyword in market_name for keyword in
                   ['3 Or More Saves', '3+ Saves', 'To Make 3', 'Make 3 Or More']):
                for selection in market.get('selections', []):
                    gk_name = selection.get('name')
                    odds = selection.get('price', {}).get('winPrice')

                    if gk_name and odds:
                        if gk_name not in gk_data:
                            gk_data[gk_name] = {'3_saves': None, '6_saves': None, 'team': None}
                        gk_data[gk_name]['3_saves'] = odds

            # Check for 6+ saves markets
            elif any(keyword in market_name for keyword in
                     ['6 Or More Saves', '6+ Saves', 'To Make 6', 'Make 6 Or More']):
                for selection in market.get('selections', []):
                    gk_name = selection.get('name')
                    odds = selection.get('price', {}).get('winPrice')

                    if gk_name and odds:
                        if gk_name not in gk_data:
                            gk_data[gk_name] = {'3_saves': None, '6_saves': None, 'team': None}
                        gk_data[gk_name]['6_saves'] = odds

        # Second pass: determine which team each goalkeeper plays for
        # This is a simplified approach - you might need to enhance this based on actual data
        participant1 = match_info.get('participant1', '')
        participant2 = match_info.get('participant2', '')

        # Simple team assignment (you may need to improve this logic based on actual data)
        gk_count = 0
        for gk_name in gk_data.keys():
            if gk_count == 0:
                gk_data[gk_name]['team'] = participant1
            else:
                gk_data[gk_name]['team'] = participant2
            gk_count += 1

        # Create final structure for each goalkeeper
        for gk_name, data in gk_data.items():
            if data['3_saves'] is not None or data['6_saves'] is not None:
                goalkeeper_saves_odds.append({
                    'Match': f"{participant1} v {participant2}",
                    'Date': match_info.get('formatted_date', ''),
                    'Team': data['team'] or 'Unknown',
                    'Goalkeeper': gk_name,
                    '3+ Saves': data['3_saves'],
                    '6+ Saves': data['6_saves']
                })

        return goalkeeper_saves_odds

    def scrape_all_odds(self, delay_seconds: float = 1.0, discover_mode: bool = False) -> tuple:
        """
        Scrape all odds for all matches across all markets
        """
        print("Fetching match IDs...")
        matches = self.get_match_ids()

        if not matches:
            print("No matches found")
            return tuple([[] for _ in range(10)])

        # Discovery mode to find market IDs
        if discover_mode:
            print("DISCOVERY MODE: Showing all markets for first match...")
            self.discover_markets(matches[0]['match_id'])
            return tuple([[] for _ in range(10)])

        # Initialize lists for all market types
        all_win_draw_win = []
        all_goalscorer_odds = []
        all_assist_odds = []
        all_team_goals_odds = []
        all_clean_sheet_odds = []
        all_booking_odds = []
        all_two_goals_odds = []
        all_two_assists_odds = []
        all_goalkeeper_saves_odds = []

        for i, match in enumerate(matches, 1):
            print(f"Processing match {i}/{len(matches)}: {match['match_name']}")

            markets_data = self.get_match_markets(match['match_id'])

            if markets_data:
                # Extract all types of odds
                win_draw_win = self.extract_win_draw_win_odds(markets_data, match)
                goalscorer_odds = self.extract_goalscorer_odds(markets_data, match)
                assist_odds = self.extract_assist_odds(markets_data, match)
                team_goals_odds = self.extract_team_goals_odds(markets_data, match)
                clean_sheet_odds = self.extract_clean_sheet_odds(markets_data, match)
                booking_odds = self.extract_to_be_booked_odds(markets_data, match)
                two_goals_odds = self.extract_two_or_more_goals_odds(markets_data, match)
                two_assists_odds = self.extract_two_or_more_assists_odds(markets_data, match)
                goalkeeper_saves_odds = self.extract_goalkeeper_saves_odds(markets_data, match)

                # Add to main lists
                all_win_draw_win.extend(win_draw_win)
                all_goalscorer_odds.extend(goalscorer_odds)
                all_assist_odds.extend(assist_odds)
                all_team_goals_odds.extend(team_goals_odds)
                all_clean_sheet_odds.extend(clean_sheet_odds)
                all_booking_odds.extend(booking_odds)
                all_two_goals_odds.extend(two_goals_odds)
                all_two_assists_odds.extend(two_assists_odds)
                all_goalkeeper_saves_odds.extend(goalkeeper_saves_odds)

                print(f"  Found: {len(win_draw_win)} WDW, {len(goalscorer_odds)} goalscorer, "
                      f"{len(assist_odds)} assist, {len(team_goals_odds)} team goals, "
                      f"{len(clean_sheet_odds)} clean sheet, {len(booking_odds)} booking, "
                      f"{len(two_goals_odds)} 2+ goals, {len(two_assists_odds)} 2+ assists, "
                      f"{len(goalkeeper_saves_odds)} GK saves")
            else:
                print(f"  No markets data found for match {match['match_id']}")

            if i < len(matches):
                time.sleep(delay_seconds)

        return (all_win_draw_win, all_goalscorer_odds, all_assist_odds, all_team_goals_odds,
                all_clean_sheet_odds, all_booking_odds, all_two_goals_odds, all_two_assists_odds,
                all_goalkeeper_saves_odds)

    def populate_excel_columns(self, excel_file, filename, dataframe, start_row=2, include_headers=True):
        """
        Populate Excel columns starting from A with dataframe data and optionally headers.

        Parameters:
        - excel_file: path to Excel file
        - sheet_name: name of the worksheet
        - dataframe: pandas DataFrame with data to insert
        - start_row: row number to start inserting data (default: 2)
        - include_headers: whether to write column names to row 1 (default: True)
        """

        sheet_name = ''

        if filename == 'sportsbet_win_draw_win_odds.csv':
            sheet_name = 'Fixture Odds'
        elif filename == 'sportsbet_goalscorer_odds.csv':
            sheet_name = 'Score'
        elif filename == 'sportsbet_assist_odds.csv':
            sheet_name = 'Assist'
        elif filename == 'sportsbet_team_goals_odds.csv':
            sheet_name = 'Team Total Goals'
        elif filename == 'sportsbet_team_goals_odds_f2.csv':
            sheet_name = 'F2 Team Total Goals'
        elif filename == 'sportsbet_clean_sheet_odds.csv':
            sheet_name = 'Clean Sheet'
        elif filename == 'sportsbet_clean_sheet_odds_f2.csv':
            sheet_name = 'F2 Clean Sheet'
        elif filename == 'sportsbet_booking_odds.csv':
            sheet_name = 'Yellow Card'
        elif filename == 'sportsbet_two_goals_odds.csv':
            sheet_name = 'Score 2+'
        elif filename == 'sportsbet_two_assists_odds.csv':
            sheet_name = 'Assist 2+'
        elif filename == 'sportsbet_goalkeeper_saves_odds.csv':
            sheet_name = 'Goalkeeper Saves'

        if sheet_name != '':

            workbook = load_workbook(excel_file)
            worksheet = workbook[sheet_name]

            num_cols = len(dataframe.columns)

            # Add headers if requested
            if include_headers:
                for col_idx, col_name in enumerate(dataframe.columns):
                    excel_col = get_column_letter(col_idx + 1)
                    worksheet[f'{excel_col}1'].value = col_name

            # Clear existing data in the data range
            for col_idx in range(num_cols):
                excel_col = get_column_letter(col_idx + 1)  # A=1, B=2, etc.
                # Clear the entire column by deleting all values from start_row to max_row
                for row in range(start_row, worksheet.max_row + 1):
                    worksheet[f'{excel_col}{row}'].value = None

            # Insert new data
            for row_idx, (_, row_data) in enumerate(dataframe.iterrows(), start=start_row):
                for col_idx, col_name in enumerate(dataframe.columns):
                    excel_col = get_column_letter(col_idx + 1)
                    value = row_data[col_name]

                    # Handle NaN values
                    if pd.isna(value):
                        worksheet[f'{excel_col}{row_idx}'].value = None
                    else:
                        worksheet[f'{excel_col}{row_idx}'].value = value

            workbook.save(excel_file)

            print(f"Successfully updated columns A-{get_column_letter(num_cols)} ({num_cols} columns, {len(dataframe)} rows) in {sheet_name}")

    def save(self, all_data: tuple):
        """
        Save all odds data to CSV files
        """
        (win_draw_win, goalscorer_odds, assist_odds, team_goals_odds,
         clean_sheet_odds, booking_odds, two_goals_odds, two_assists_odds,
         goalkeeper_saves_odds) = all_data

        files_saved = []

        # Save each market type to separate CSV files
        datasets = [
            (win_draw_win, 'sportsbet_win_draw_win_odds.csv'),
            (goalscorer_odds, 'sportsbet_goalscorer_odds.csv'),
            (assist_odds, 'sportsbet_assist_odds.csv'),
            (team_goals_odds, 'sportsbet_team_goals_odds.csv'),
            (team_goals_odds, 'sportsbet_team_goals_odds_f2.csv'),
            (clean_sheet_odds, 'sportsbet_clean_sheet_odds.csv'),
            (clean_sheet_odds, 'sportsbet_clean_sheet_odds_f2.csv'),
            (booking_odds, 'sportsbet_booking_odds.csv'),
            (two_goals_odds, 'sportsbet_two_goals_odds.csv'),
            (two_assists_odds, 'sportsbet_two_assists_odds.csv'),
            (goalkeeper_saves_odds, 'sportsbet_goalkeeper_saves_odds.csv')
        ]

        for data, filename in datasets:
            if data:
                df = pd.DataFrame(data)

                if filename in ['sportsbet_team_goals_odds.csv', 'sportsbet_clean_sheet_odds.csv']:
                    df = df.head(20)
                elif filename in ['sportsbet_team_goals_odds_f2.csv', 'sportsbet_clean_sheet_odds_f2.csv']:
                    df = df.tail(20)

                df.to_csv(filename, index=False)
                files_saved.append(f"{filename} ({len(df)} rows)")

                self.populate_excel_columns('Fantasy Premier League.xlsx', filename, df)


        print("\nFiles saved:")
        for file_info in files_saved:
            print(f"  - {file_info}")

        if not files_saved:
            print("No data to save")


def main():
    """
    Main function to run the comprehensive scraper
    """
    scraper = ComprehensiveSportsbetScraper()

    print("Starting Sportsbet Odds Scraping...")
    print("-" * 80)

    # Option to discover markets first
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "discover":
        print("Running in discovery mode to find market IDs...")
        scraper.scrape_all_odds(discover_mode=True)
        return

    # Scrape all odds
    all_data = scraper.scrape_all_odds(delay_seconds=2.0)

    print("-" * 80)
    print("Scraping complete!")

    # Print summary
    market_names = ['Win-Draw-Win', 'Goalscorer', 'Assist', 'Team Goals', 'Clean Sheet', 'Booking',
                    '2+ Goals', '2+ Assists', 'Goalkeeper Saves']

    for i, (data, name) in enumerate(zip(all_data, market_names)):
        print(f"Total {name} odds collected: {len(data)}")

    # Save to CSV files
    scraper.save(all_data)

    print("\nFiles saved successfully!")
    print("\nTo discover available markets, run: python script.py discover")


if __name__ == "__main__":
    main()
